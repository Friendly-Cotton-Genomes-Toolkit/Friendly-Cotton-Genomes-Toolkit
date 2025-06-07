import pandas as pd
from scripts.EasyOpenAI import EasyOpenAIWrapper
import json
import re
from collections import Counter
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from typing import List, Tuple, Optional, Any, Dict

# --- Caching Imports (借鉴自 universal_openai_request.py) ---
from diskcache import Cache
import hashlib
import time

# --- Progress Bar Import ---
from tqdm import tqdm  # 导入tqdm


# --- 配置参数 ---
DEFAULT_SHEET_NAME_MAIN: str = '拟南芥结果汇总'
DEFAULT_SHEET_NAME_KEYWORDS_COTTON: str = '棉花关键词'
DEFAULT_SHEET_NAME_FREQUENCY_ANALYSIS: str = '频率分析'
DEFAULT_COLUMN_GENE_FUNCTION: str = '功能解释'
DEFAULT_COLUMN_ARABIDOPSIS_KEYWORDS: str = '拟南芥关键词'
DEFAULT_PLOT_FILENAME: str = 'keyword_frequency_plot.png'
DEFAULT_MAX_WORKERS_OPENAI: int = 12
DEFAULT_TASK_IDENTIFIER: str = "arabidopsis_keyword_extraction"
CACHE_DIRECTORY_BASE_ARABIDOPSIS: str = "tmp/.arabidopsis_cache"
PROCESSING_ERROR_MARKER_ARABIDOPSIS: str = "---PROCESSING_ERROR---"


# --- 辅助函数 ---
def _find_matplotlib_font() -> Optional[str]:
    font_paths: List[str] = fm.findSystemFonts(fontpaths=None, fontext='ttf')
    preferred_fonts: List[str] = ['simhei', 'microsoft yahei', 'wqy', 'source han sans', 'arial unicode ms', 'msyh']
    for font_path in font_paths:
        try:
            font_name: str = fm.FontProperties(fname=font_path).get_name()
            if any(keyword in font_name.lower() for keyword in preferred_fonts):
                return font_name
        except RuntimeError:
            continue
    if font_paths:
        try:
            return fm.FontProperties(fname=font_paths[0]).get_name()
        except RuntimeError:
            pass
    return None


CHINESE_FONT: Optional[str] = _find_matplotlib_font()
if CHINESE_FONT:
    plt.rcParams['font.sans-serif'] = [CHINESE_FONT]
    plt.rcParams['axes.unicode_minus'] = False
    print(f"已设置 Matplotlib 字体为: {CHINESE_FONT}")
else:
    print(
        "警告: 未找到合适的中文TTF字体，绘图中的中文可能无法正常显示。请安装支持中文的字体（如SimHei, Microsoft YaHei等）并确保Matplotlib可以找到它。")


# --- 核心OpenAI调用函数 (带缓存和重试) ---
def _get_keywords_from_openai_with_cache(
        function_description: str,
        reference_keywords_str: str,
        client: EasyOpenAIWrapper,
        cache: Cache,
        task_identifier: str,
        retries: int = 3,
        delay: int = 5
) -> List[str]:
    if not function_description or pd.isna(function_description):
        return []

    if not client or not hasattr(client, 'api_key') or not client.api_key:
        print(f"错误 ({task_identifier}): OpenAI API Key 未在客户端中有效设置。描述: {function_description[:30]}...")
        return []

    desc_hash = hashlib.md5(str(function_description).encode('utf-8')).hexdigest()
    ref_hash = hashlib.md5(reference_keywords_str.encode('utf-8')).hexdigest()
    cache_key = f"{task_identifier}::desc_hash:{desc_hash}::ref_hash:{ref_hash}"

    cached_result = cache.get(cache_key)
    if cached_result is not None:
        if cached_result == PROCESSING_ERROR_MARKER_ARABIDOPSIS:
            return []
        return cached_result

    prompt = f"""
    请分析以下基因功能解释，并参考提供的关键词列表，总结出与功能解释最相关的核心关键词。
    返回的关键词应该能准确概括其主要功能，可以有多个关键词。
    请以JSON列表的格式返回结果，例如：["关键词1", "关键词2", "关键词3"]。

    基因功能解释：
    "{function_description}"

    参考关键词列表（请学习这些词的风格和领域，但不要直接从中选择，除非它们也明确适用于上述功能解释）：
    "{reference_keywords_str}"

    你的关键词总结（JSON列表格式）：
    """
    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                messages=[
                    {"role": "system", "content": "你是一个生物信息学助手，擅长从文本中提取核心关键词。"},
                    {"role": "user", "content": prompt}
                ],
            )
            content: str = response.choices[0].message.content
            match = re.search(r'```json\s*([\s\S]*?)\s*```', content)
            json_str: str = match.group(1) if match else content

            try:
                keywords_data: Any = json.loads(json_str)
                if isinstance(keywords_data, list) and all(isinstance(kw, str) for kw in keywords_data):
                    extracted_keywords = [kw.strip() for kw in keywords_data if kw.strip()]
                    cache.set(cache_key, extracted_keywords)
                    return extracted_keywords
                elif isinstance(keywords_data, str):
                    extracted_keywords = [kw.strip() for kw in re.split(r'[、,;]', keywords_data) if kw.strip()]
                    cache.set(cache_key, extracted_keywords)
                    return extracted_keywords
                else:
                    print(
                        f"警告 ({task_identifier}): OpenAI返回数据格式非预期: {keywords_data}。描述: {function_description[:30]}...")
            except json.JSONDecodeError:
                cleaned_content: str = re.sub(r'[\[\]"\']', '', json_str).strip()
                potential_keywords: List[str] = [
                    kw.strip() for kw in re.split(r'[、,;]', cleaned_content) if kw.strip()
                ]
                if potential_keywords:
                    print(
                        f"提示 ({task_identifier}): JSON解码失败，但尝试从文本中提取: {potential_keywords}。描述: {function_description[:30]}...")
                    cache.set(cache_key, potential_keywords)
                    return potential_keywords
                print(
                    f"JSON解码错误 ({task_identifier}) 无法从文本修复。内容: {content}。描述: {function_description[:30]}...")

            print(
                f"API调用成功但响应解析失败 ({task_identifier}) - 尝试 {attempt + 1}/{retries}。描述: {function_description[:30]}...")
            if attempt < retries - 1:
                time.sleep(delay * (attempt + 1))
            else:
                break
        except Exception as e:
            print(
                f"API调用时发生错误 ({task_identifier}): {e}。尝试 {attempt + 1}/{retries}。描述: {function_description[:30]}...")
            if attempt < retries - 1:
                time.sleep(delay * (attempt + 1))
            else:
                break
    print(f"警告 ({task_identifier}): 经过 {retries} 次尝试后，处理描述 '{function_description[:30]}...' 失败。")
    cache.set(cache_key, PROCESSING_ERROR_MARKER_ARABIDOPSIS)
    return []


def _process_row_for_openai(args: Tuple[int, pd.Series, str, str, EasyOpenAIWrapper, Cache, str]) -> Tuple[
    int, List[str]]:
    index, row_data, column_gene_function, reference_keywords_str, client_instance, cache_instance, task_id = args
    function_description: str = str(row_data[column_gene_function])
    keywords: List[str] = _get_keywords_from_openai_with_cache(
        function_description, reference_keywords_str, client_instance, cache_instance, task_id
    )
    return index, keywords


# --- 主入口函数 ---
def analyze_excel_data(
        excel_file_path: str,
        sheet_name_main: str,
        sheet_name_keywords_cotton: str,
        sheet_name_frequency_analysis: str,
        column_gene_function: str,
        column_arabidopsis_keywords: str,
        client: EasyOpenAIWrapper,
        task_identifier: str = DEFAULT_TASK_IDENTIFIER,
        max_workers_openai: int = DEFAULT_MAX_WORKERS_OPENAI,
        plot_filename_base: str = DEFAULT_PLOT_FILENAME
) -> bool:
    print(f"开始处理Excel文件: {excel_file_path} (任务: {task_identifier})")
    plot_actual_filename: Optional[str] = None

    current_cache_dir = f"{CACHE_DIRECTORY_BASE_ARABIDOPSIS}_{task_identifier}"
    os.makedirs(current_cache_dir, exist_ok=True)
    cache = Cache(current_cache_dir)
    print(f"任务 '{task_identifier}' 的缓存将使用目录: {os.path.abspath(current_cache_dir)}")

    try:
        try:
            xls: pd.ExcelFile = pd.ExcelFile(excel_file_path)
        except FileNotFoundError:
            print(f"错误: Excel文件未找到于路径 '{excel_file_path}'。")
            return False
        except Exception as e:
            print(f"读取Excel文件 '{excel_file_path}' 时发生错误: {e}")
            return False

        if sheet_name_main not in xls.sheet_names:
            print(f"错误: 工作表 '{sheet_name_main}' 在Excel文件中未找到。")
            return False
        df_main: pd.DataFrame = pd.read_excel(xls, sheet_name=sheet_name_main)
        print(f"成功读取工作表: {sheet_name_main}，包含 {len(df_main)} 行数据。")

        if column_gene_function not in df_main.columns:
            print(f"错误: 在工作表 '{sheet_name_main}' 中未找到表头 '{column_gene_function}'。")
            return False

        if sheet_name_keywords_cotton not in xls.sheet_names:
            print(f"错误: 工作表 '{sheet_name_keywords_cotton}' 在Excel文件中未找到。")
            return False
        df_cotton_keywords: pd.DataFrame = pd.read_excel(xls, sheet_name=sheet_name_keywords_cotton)
        if '关键词' not in df_cotton_keywords.columns:
            print(f"错误: 工作表 '{sheet_name_keywords_cotton}' 中未找到表头 '关键词'。")
            return False
        reference_keywords: List[str] = df_cotton_keywords['关键词'].dropna().astype(str).tolist()
        reference_keywords_str: str = "、".join(reference_keywords)
        print(f"成功读取参考关键词: {len(reference_keywords)} 个。")

        if column_arabidopsis_keywords in df_main.columns:
            print(f"发现已存在的列 '{column_arabidopsis_keywords}'，正在删除...")
            df_main = df_main.drop(columns=[column_arabidopsis_keywords])

        print(f"开始使用OpenAI提取关键词 (任务: {task_identifier}, 最多 {max_workers_openai} 个线程)...")
        all_generated_keywords_for_column: List[List[str]] = [[] for _ in range(len(df_main))]

        tasks_args_list: List[Tuple[int, pd.Series, str, str, EasyOpenAIWrapper, Cache, str]] = []
        for index, row in df_main.iterrows():
            tasks_args_list.append((
                int(index), row, column_gene_function, reference_keywords_str,
                client, cache, task_identifier
            ))

        # 使用ThreadPoolExecutor和tqdm显示进度条
        with ThreadPoolExecutor(max_workers=max_workers_openai) as executor:
            futures: List[Any] = [executor.submit(_process_row_for_openai, args) for args in tasks_args_list]
            # 使用tqdm包装as_completed来显示进度
            for i, future in enumerate(
                    tqdm(as_completed(futures), total=len(futures), desc=f"处理基因功能 ({task_identifier})", unit="行",
                         ncols=100, leave=False)):
                try:
                    idx, keywords_list = future.result()
                    all_generated_keywords_for_column[idx] = keywords_list
                except Exception as e:
                    original_args_index = -1
                    # 查找与当前future关联的原始任务参数，以获取索引
                    # 注意: 这个查找可能不是最高效的，但对于错误报告是可行的
                    for k_idx, f_check in enumerate(futures):  # Iterate over the original list of futures
                        if f_check == future:
                            # tasks_args_list[k_idx] 是原始参数元组
                            # tasks_args_list[k_idx][0] 是原始DataFrame的索引
                            original_args_index = tasks_args_list[k_idx][0]
                            break
                    print(
                        f"处理行 (原始索引 {original_args_index}, 任务 {i + 1}/{len(futures)}) 时发生错误: {e} (任务: {task_identifier})")

        df_main[column_arabidopsis_keywords] = ["、".join(kws) if kws else "" for kws in
                                                all_generated_keywords_for_column]
        print(f"\n已将生成的关键词添加到新列 '{column_arabidopsis_keywords}'。")  # 添加换行符以避免与tqdm重叠

        if column_gene_function in df_main.columns:
            function_col_idx: int = df_main.columns.get_loc(column_gene_function)
            cols: List[str] = list(df_main.columns)
            if column_arabidopsis_keywords in cols:
                cols.remove(column_arabidopsis_keywords)
                cols.insert(function_col_idx + 1, column_arabidopsis_keywords)
                df_main = df_main[cols]
                print(f"已将 '{column_arabidopsis_keywords}' 列移动到 '{column_gene_function}' 列之后。")

        print("开始分析关键词频率...")
        all_individual_keywords: List[str] = []
        for keywords_str_val in df_main[column_arabidopsis_keywords].dropna().astype(str):
            if keywords_str_val:
                all_individual_keywords.extend([kw.strip() for kw in keywords_str_val.split('、') if kw.strip()])
        keyword_counts: Counter = Counter(all_individual_keywords)
        df_frequency: pd.DataFrame = pd.DataFrame(
            keyword_counts.items(), columns=['关键词', '频率']
        ).sort_values(by='频率', ascending=False)
        print("关键词频率计算完成。")

        print("开始绘制关键词频率图...")
        plt.figure(figsize=(15, 10))
        top_n: int = min(30, len(df_frequency))
        if top_n > 0:
            bars = plt.bar(df_frequency['关键词'][:top_n], df_frequency['频率'][:top_n], color='steelblue')
            plt.xlabel('关键词', fontsize=14, labelpad=10)
            plt.ylabel('频率', fontsize=14, labelpad=10)
            plt.title(f'拟南芥关键词频率分析 (Top {top_n}) - 任务: {task_identifier}', fontsize=18, pad=20)
            plt.xticks(rotation=60, ha="right", fontsize=10)
            plt.yticks(fontsize=10)
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            for bar in bars:
                yval = bar.get_height()
                plt.text(bar.get_x() + bar.get_width() / 2.0, yval + 0.05 * df_frequency['频率'][:top_n].max(),
                         int(yval) if yval == int(yval) else round(yval, 1),
                         ha='center', va='bottom', fontsize=8)
            plt.tight_layout()
            plot_actual_filename = os.path.join(os.getcwd(),
                                                f"{os.path.splitext(plot_filename_base)[0]}_{task_identifier}{os.path.splitext(plot_filename_base)[1]}")
            plt.savefig(plot_actual_filename, dpi=300, bbox_inches='tight')
            print(f"关键词频率图已保存为: {plot_actual_filename}")
        else:
            print("没有关键词可供绘图。")
            plot_actual_filename = None

        print(f"准备将结果写入Excel文件 '{excel_file_path}'...")
        try:
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='w') as writer:
                df_main.to_excel(writer, sheet_name=sheet_name_main, index=False)
                print(f"已写入主数据到工作表: '{sheet_name_main}'")
                df_frequency.to_excel(writer, sheet_name=sheet_name_frequency_analysis, index=False, startrow=0)
                print(f"已写入频率分析到工作表: '{sheet_name_frequency_analysis}'")

            if plot_actual_filename and os.path.exists(plot_actual_filename):
                book_final = load_workbook(excel_file_path)
                if sheet_name_frequency_analysis in book_final.sheetnames:
                    ws_freq = book_final[sheet_name_frequency_analysis]
                    img = OpenpyxlImage(plot_actual_filename)
                    img.anchor = f'A{len(df_frequency) + 3}'
                    ws_freq.add_image(img)
                    print(f"已将频率图添加到 '{sheet_name_frequency_analysis}' 工作表。")
                    book_final.save(excel_file_path)
                else:
                    print(f"警告: 工作表 '{sheet_name_frequency_analysis}' 在pandas写入后未找到，无法添加图片。")
            elif plot_actual_filename:
                print(f"警告: 频率图文件 '{plot_actual_filename}' 未找到，无法添加到Excel。")
            else:
                print("提示: 没有频率图文件需要添加到Excel。")
            print(f"处理完成！结果已保存到 '{excel_file_path}'。")
            return True
        except PermissionError:
            print(f"错误: 权限不足，无法写入Excel文件 '{excel_file_path}'。请确保文件未被打开或具有写入权限。")
            return False
        except Exception as e:
            print(f"写入Excel或添加图片时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False
    finally:
        if plot_actual_filename and os.path.exists(plot_actual_filename):
            try:
                os.remove(plot_actual_filename)
                print(f"已删除临时图片文件: {plot_actual_filename}")
            except Exception as e_remove:
                print(f"删除临时图片文件 '{plot_actual_filename}' 失败: {e_remove}")
        plt.close('all')
        if 'cache' in locals() and cache is not None:
            cache.close()
            print(f"任务 '{task_identifier}' 的磁盘缓存已关闭。")


# --- 如何调用这个入口函数 ---
if __name__ == '__main__':
    excel_path: str = 'your_excel_file.xlsx'  # <--- 修改这里
    my_openai_key: str = 'sk-YOUR_API_KEY_HERE'  # <--- 重要: 替换为你的真实API密钥
    # 或者使用 'mock' 来测试流程而不产生费用
    # my_openai_key: str = 'mock'

    # 示例文件创建逻辑
    if not os.path.exists(excel_path):
        print(f"示例Excel文件 '{excel_path}' 不存在，将创建一个...")
        try:
            data_main_example: Dict[str, List[Any]] = {
                'gene': [f'Gene{i + 1}' for i in range(10)],
                '功能解释': [
                    "参与植物生长素信号转导途径，调节根系发育和顶端优势。",
                    "编码一个关键的酶，催化光合作用碳固定过程中的一个重要步骤。",
                    "在植物受到病原菌侵染时被诱导表达，可能参与植物的免疫防御反应。",
                    "与细胞壁的合成和修饰有关，影响植物细胞的形态和机械强度。",
                    "调控开花时间，响应环境中的光周期和温度变化。",
                    "一个转录因子，结合到特定DNA序列，激活或抑制下游基因的表达，参与胁迫应答。",
                    "编码离子通道蛋白，负责特定离子（如钾离子、钙离子）的跨膜运输。",
                    "参与植物次生代谢产物（如花青素、生物碱）的生物合成途径。",
                    "在种子萌发和休眠调控中起作用，可能与脱落酸或赤霉素的平衡有关。", "编码热激蛋白，在高温胁迫下保护细胞内其他蛋白质免于变性失活。"
                ]
            }  # 省略其他列以保持简洁
            df_main_example_pd: pd.DataFrame = pd.DataFrame(data_main_example)
            data_cotton_example: Dict[str, List[str]] = {
                '关键词': ['生长发育', '胁迫响应', '信号通路', '代谢调控', '转录调控', '光合作用', '细胞结构',
                           '激素平衡', '离子运输', '开花']
            }
            df_cotton_example_pd: pd.DataFrame = pd.DataFrame(data_cotton_example)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer_example:
                df_main_example_pd.to_excel(writer_example, sheet_name=DEFAULT_SHEET_NAME_MAIN, index=False)
                df_cotton_example_pd.to_excel(writer_example, sheet_name=DEFAULT_SHEET_NAME_KEYWORDS_COTTON,
                                              index=False)
            print(f"已创建示例Excel文件: '{excel_path}'。")
        except Exception as e_create:
            print(f"创建示例Excel文件失败: {e_create}")
            exit()

    actual_client: EasyOpenAIWrapper
    current_task_id = "arabidopsis_run_002"  # 可以修改或参数化这个ID

    if my_openai_key == 'sk-YOUR_API_KEY_HERE' or not my_openai_key:
        print("错误：请在脚本中设置你的 OpenAI API Key (变量 my_openai_key)。")
        print("你可以设置为 'mock' 来使用模拟客户端进行测试。")
        exit()
    elif my_openai_key == 'mock':
        print("信息: 使用模拟模式运行。")
        actual_client = EasyOpenAIWrapper(api_key=None, preset='mock')
    else:
        print(f"准备使用 API Key: {my_openai_key[:8]}...{my_openai_key[-4:]}")
        actual_client = EasyOpenAIWrapper(api_key=my_openai_key, preset='gemini')

    success: bool = analyze_excel_data(
        excel_file_path=excel_path,
        sheet_name_main=DEFAULT_SHEET_NAME_MAIN,
        sheet_name_keywords_cotton=DEFAULT_SHEET_NAME_KEYWORDS_COTTON,
        sheet_name_frequency_analysis=DEFAULT_SHEET_NAME_FREQUENCY_ANALYSIS,
        column_gene_function=DEFAULT_COLUMN_GENE_FUNCTION,
        column_arabidopsis_keywords=DEFAULT_COLUMN_ARABIDOPSIS_KEYWORDS,
        client=actual_client,
        task_identifier=current_task_id
    )
    if success:
        print(f"Excel分析流程执行成功 (任务: {current_task_id})。结果已保存至 '{excel_path}'")
    else:
        print(f"Excel分析流程执行失败或遇到错误 (任务: {current_task_id})。请检查上述日志。")